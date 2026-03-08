#!/usr/bin/env python3
"""
Detailed Extraction - Key Sheets Analysis
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import re

def extract_cell_info(cell):
    if cell.data_type == 'f':
        return {'type': 'formula', 'content': str(cell.value)}
    elif cell.value is not None:
        return {'type': 'value', 'content': cell.value}
    return None

def extract_named_tables(ws, sheet_name):
    """Try to identify tables/regions in a sheet"""
    tables = []
    
    # Look for headers
    for row in range(1, min(ws.max_row, 20)):
        for col in range(1, min(ws.max_column, 20)):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                if any(keyword in cell.value.lower() for keyword in ['tableau', 'tarif', 'prix', 'cout', 'parametre']):
                    tables.append({
                        'name': cell.value,
                        'start_row': row,
                        'start_col': col
                    })
    
    return tables

def main():
    file_path = r'C:\Users\lenovo\Desktop\printpilot\docs\Tableau Numérique-Offset HAVET-IMB 24-02-2026 (1).xlsm'
    output_dir = r'C:\Users\lenovo\Desktop\printpilot\docs\Extraction'
    
    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
    
    # Focus on key sheets
    key_sheets = [
        'Donnees Num',
        'Tableau papier',
        'Tableau Façonnage Num',
        'Tableau Façonnage OFFSET',
        'Livraison',
        'Transports',
        'Parc Machines IMB',
        'Réglages Presses'
    ]
    
    all_formulas = []
    all_variables = []
    
    for sheet_name in key_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet not found: {sheet_name}")
            continue
            
        print(f"Analyzing: {sheet_name}...")
        ws = wb[sheet_name]
        
        sheet_data = {
            'name': sheet_name,
            'formulas': [],
            'variables': [],
            'tables': []
        }
        
        # Extract all formulas and variables
        for row in range(1, min(ws.max_row + 1, 300)):
            for col in range(1, min(ws.max_column + 1, 50)):
                cell = ws.cell(row=row, column=col)
                cell_ref = f'{get_column_letter(col)}{row}'
                
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    sheet_data['formulas'].append({
                        'cell': cell_ref,
                        'formula': formula
                    })
                    all_formulas.append({
                        'sheet': sheet_name,
                        'cell': cell_ref,
                        'formula': formula
                    })
                elif cell.value is not None:
                    # Look for variable-like cells (labels with values nearby)
                    if isinstance(cell.value, str) and len(cell.value) > 2:
                        # Check if next cell has a value
                        next_cell = ws.cell(row=row, column=col+1)
                        if next_cell.value is not None and not isinstance(next_cell.value, str):
                            sheet_data['variables'].append({
                                'name': cell.value,
                                'cell': cell_ref,
                                'value_cell': f'{get_column_letter(col+1)}{row}',
                                'value': next_cell.value,
                                'is_formula': next_cell.data_type == 'f'
                            })
                            all_variables.append({
                                'sheet': sheet_name,
                                'name': cell.value,
                                'value': next_cell.value,
                                'cell': f'{get_column_letter(col+1)}{row}'
                            })
        
        # Save individual sheet analysis
        sheet_file = os.path.join(output_dir, f"{sheet_name.replace(' ', '_')}_analysis.json")
        with open(sheet_file, 'w', encoding='utf-8') as f:
            json.dump(sheet_data, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"  Found {len(sheet_data['formulas'])} formulas, {len(sheet_data['variables'])} variables")
    
    # Save all formulas
    formulas_file = os.path.join(output_dir, 'all_formulas.json')
    with open(formulas_file, 'w', encoding='utf-8') as f:
        json.dump(all_formulas, f, indent=2, ensure_ascii=False)
    
    # Save all variables
    variables_file = os.path.join(output_dir, 'all_variables.json')
    with open(variables_file, 'w', encoding='utf-8') as f:
        json.dump(all_variables, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\nExtraction complete!")
    print(f"Total formulas: {len(all_formulas)}")
    print(f"Total variables: {len(all_variables)}")
    
    wb.close()

if __name__ == '__main__':
    main()
