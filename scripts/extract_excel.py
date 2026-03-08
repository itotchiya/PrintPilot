#!/usr/bin/env python3
"""
Excel Extraction Script for PrintPilot
Extracts all data from the HAVET-IMB pricing Excel file
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import os

def extract_cell_info(cell):
    """Extract value or formula from a cell"""
    if cell.data_type == 'f':
        return {'type': 'formula', 'content': cell.value}
    elif cell.value is not None:
        return {'type': 'value', 'content': cell.value}
    return None

def main():
    file_path = r'C:\Users\lenovo\Desktop\printpilot\docs\Tableau Numérique-Offset HAVET-IMB 24-02-2026 (1).xlsm'
    output_dir = r'C:\Users\lenovo\Desktop\printpilot\docs\Extraction'
    
    print(f"Loading: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
    
    print(f"Found {len(wb.sheetnames)} sheets:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {name}")
    
    # Summary data
    summary = {}
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        print(f"\nProcessing: {sheet_name}...")
        ws = wb[sheet_name]
        
        max_row = min(ws.max_row, 200)  # Limit for initial extraction
        max_col = min(ws.max_column, 60)
        
        sheet_data = {
            'name': sheet_name,
            'total_rows': ws.max_row,
            'total_cols': ws.max_column,
            'extracted_rows': max_row,
            'extracted_cols': max_col,
            'cells': []
        }
        
        # Extract all cells with values or formulas
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                info = extract_cell_info(cell)
                if info:
                    info['cell'] = f'{get_column_letter(col)}{row}'
                    info['row'] = row
                    info['col'] = col
                    sheet_data['cells'].append(info)
        
        summary[sheet_name] = sheet_data
        print(f"  Extracted {len(sheet_data['cells'])} cells with data")
    
    # Save summary JSON
    summary_file = os.path.join(output_dir, 'raw_extraction.json')
    with open(summary_file, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n✓ Raw extraction saved to: {summary_file}")
    
    # Create markdown summary
    md_file = os.path.join(output_dir, '00-Sheets-Overview.md')
    with open(md_file, 'w', encoding='utf-8') as f:
        f.write("# Excel Sheets Overview\n\n")
        f.write(f"**Source File:** `Tableau Numérique-Offset HAVET-IMB 24-02-2026 (1).xlsm`\n\n")
        f.write(f"**Total Sheets:** {len(wb.sheetnames)}\n\n")
        f.write("## Sheet List\n\n")
        f.write("| # | Sheet Name | Rows | Cols | Description |\n")
        f.write("|---|------------|------|------|-------------|\n")
        
        for i, (name, data) in enumerate(summary.items(), 1):
            desc = get_sheet_description(name)
            f.write(f"| {i} | `{name}` | {data['total_rows']} | {data['total_cols']} | {desc} |\n")
    
    print(f"✓ Overview saved to: {md_file}")
    wb.close()

def get_sheet_description(name):
    """Get description based on sheet name"""
    descriptions = {
        'Tableau NUMERIQUE': 'Digital printing pricing table',
        'Tableau NUMERIQUE (2)': 'Digital printing pricing (copy)',
        'Donnees Num': 'Digital data/parameters',
        'Tableau papier': 'Paper specifications and prices',
        'Livraison': 'Delivery/shipping calculations',
        'Tableau Façonnage Num': 'Digital finishing pricing',
        'Tableau OFFSET': 'Offset printing pricing table',
        'Tableau OFFSET (2)': 'Offset printing pricing (copy)',
        'Détails PRIX': 'Price breakdown details',
        'Détails PRIX DEPLIANTS': 'Leaflet price breakdown',
        'GENERATEUR BROCHURE': 'Brochure quote generator',
        'GENERATEUR FLYER-DEPLIANT': 'Flyer/leaflet quote generator',
        'Tableau Façonnage OFFSET': 'Offset finishing pricing',
        'Transports': 'Transport rates by department',
        'Bloc note': 'Notes/documentation',
        'Parc Machines IMB': 'Machine park specifications',
        'Réglages Presses': 'Press settings',
        'Sauvegarde formules': 'Formula backup'
    }
    return descriptions.get(name, 'Data sheet')

if __name__ == '__main__':
    main()
