#!/usr/bin/env python3
"""
Excel Extraction Script for PrintPilot - FIXED VERSION
"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import os

def extract_cell_info(cell):
    """Extract value or formula from a cell"""
    if cell.data_type == 'f':
        return {'type': 'formula', 'content': str(cell.value)}
    elif cell.value is not None:
        return {'type': 'value', 'content': cell.value}
    return None

def main():
    file_path = r'C:\Users\lenovo\Desktop\printpilot\docs\Tableau Numérique-Offset HAVET-IMB 24-02-2026 (1).xlsm'
    output_dir = r'C:\Users\lenovo\Desktop\printpilot\docs\Extraction'
    
    print(f"Loading: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
    
    print(f"Found {len(wb.sheetnames)} sheets")
    
    # Summary data
    summary = {}
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        print(f"Processing: {sheet_name}...")
        ws = wb[sheet_name]
        
        max_row = min(ws.max_row, 200)
        max_col = min(ws.max_column, 60)
        
        sheet_data = {
            'name': sheet_name,
            'total_rows': ws.max_row,
            'total_cols': ws.max_column,
            'extracted_rows': max_row,
            'extracted_cols': max_col,
            'cells': []
        }
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                info = extract_cell_info(cell)
                if info:
                    info['cell'] = f'{get_column_letter(col)}{row}'
                    info['row'] = row
                    info['col'] = col
                    sheet_data['cells'].append(info)
        
        summary[sheet_name] = sheet_data
        print(f"  Extracted {len(sheet_data['cells'])} cells")
    
    # Save summary JSON
    summary_file = os.path.join(output_dir, 'raw_extraction.json')
    with open(summary_file, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"Raw extraction saved to: {summary_file}")
    
    # Create markdown summary
    md_file = os.path.join(output_dir, '00-Sheets-Overview.md')
    with open(md_file, 'w', encoding='utf-8') as f:
        f.write("# Excel Sheets Overview\n\n")
        f.write("**Source File:** `Tableau Numérique-Offset HAVET-IMB 24-02-2026 (1).xlsm`\n\n")
        f.write(f"**Total Sheets:** {len(wb.sheetnames)}\n\n")
        f.write("## Sheet List\n\n")
        f.write("| # | Sheet Name | Rows | Cols | Description |\n")
        f.write("|---|------------|------|------|-------------|\n")
        
        for i, (name, data) in enumerate(summary.items(), 1):
            desc = get_sheet_description(name)
            f.write(f"| {i} | `{name}` | {data['total_rows']} | {data['total_cols']} | {desc} |\n")
        
        f.write("\n## Sheet Descriptions\n\n")
        for name in wb.sheetnames:
            desc = get_sheet_description(name)
            f.write(f"### {name}\n{desc}\n\n")
    
    print(f"Overview saved to: {md_file}")
    wb.close()

def get_sheet_description(name):
    descriptions = {
        'Tableau NUMERIQUE': 'Main digital printing pricing calculation table',
        'Tableau NUMERIQUE (2)': 'Copy of digital pricing table',
        'Donnees Num': 'Digital printing configuration data and parameters',
        'Tableau papier': 'Paper catalog with grammages and prices',
        'Livraison': 'Delivery and shipping cost calculations',
        'Tableau Façonnage Num': 'Digital finishing operations pricing',
        'Tableau OFFSET': 'Main offset printing pricing calculation table',
        'Tableau OFFSET (2)': 'Copy of offset pricing table',
        'Détails PRIX': 'Detailed price breakdown for brochures',
        'Détails PRIX DEPLIANTS': 'Detailed price breakdown for leaflets',
        'GENERATEUR BROCHURE': 'Brochure quote generator interface',
        'GENERATEUR FLYER-DEPLIANT': 'Flyer and leaflet quote generator',
        'Tableau Façonnage OFFSET': 'Offset finishing operations pricing',
        'Transports': 'Transport rates by department (French postal codes)',
        'Bloc note': 'Documentation and notes',
        'Parc Machines IMB': 'Machine park specifications',
        'Réglages Presses': 'Press settings and calibrations',
        'Sauvegarde formules': 'Backup of formulas (appears empty)'
    }
    return descriptions.get(name, 'Data sheet')

if __name__ == '__main__':
    main()
