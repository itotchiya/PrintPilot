"""Stage 4: Extract Détails PRIX (the main offset calculation detail sheet)
and the Tableau OFFSET sheet."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

XLSM = r"c:\Users\lenovo\Desktop\printpilot\Tableau Numérique-Offset HAVET-IMB 24-02-2026.xlsm"

wb = openpyxl.load_workbook(XLSM, data_only=True, keep_vba=True, read_only=True)

# Extract Details PRIX (rows 1-80, cols 1-30)
print("===== DETAILS PRIX (offset brochure calculation) =====")
ws = wb['D\u00e9tails PRIX']
for r in range(1, 80):
    vals = []
    for c in range(1, 30):
        try:
            v = ws.cell(row=r, column=c).value
        except:
            v = None
        if v is not None:
            s = str(v)
            if len(s) > 80:
                s = s[:80] + "..."
            vals.append(f"C{c}={s}")
    if vals:
        print(f"  R{r}: {' | '.join(vals)}")

# Extract Tableau OFFSET (first 30 rows for the brochure section)
print(f"\n===== TABLEAU OFFSET (main sheet) =====")
ws2 = wb['Tableau OFFSET']
for r in range(1, 35):
    vals = []
    for c in range(1, 30):
        try:
            v = ws2.cell(row=r, column=c).value
        except:
            v = None
        if v is not None:
            s = str(v)
            if len(s) > 80:
                s = s[:80] + "..."
            vals.append(f"C{c}={s}")
    if vals:
        print(f"  R{r}: {' | '.join(vals)}")

# Also extract Donnees Num for digital click rates
print(f"\n===== DONNEES NUM =====")
ws3 = wb['Donnees Num']
for r in range(1, 40):
    vals = []
    for c in range(1, 20):
        try:
            v = ws3.cell(row=r, column=c).value
        except:
            v = None
        if v is not None:
            s = str(v)
            if len(s) > 80:
                s = s[:80] + "..."
            vals.append(f"C{c}={s}")
    if vals:
        print(f"  R{r}: {' | '.join(vals)}")

wb.close()
